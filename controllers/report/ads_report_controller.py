from flask import request, jsonify, make_response
from controllers.base_controller import BaseController
from models.ads_report import AdsReport
from io import BytesIO
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class AdsReportController(BaseController):
    def post(self):
        try:
            data = request.get_json()
            ad_ids = data.get('ad_ids', [])
            start_date = data.get('start_date')
            end_date = data.get('end_date')

            # Validate required fields
            if not ad_ids:
                return jsonify({
                    "success": False,
                    "error": "Vui lòng chọn ít nhất một tin quảng cáo"
                }), 400

            if not start_date or not end_date:
                return jsonify({
                    "success": False,
                    "error": "Vui lòng chọn khoảng thời gian báo cáo"
                }), 400

            # Validate dates
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')

                if start_date_obj > end_date_obj:
                    return jsonify({
                        "success": False,
                        "error": "Ngày bắt đầu không thể sau ngày kết thúc"
                    }), 400

                # Format dates for display
                start_date_display = start_date_obj.strftime('%d/%m/%Y')
                end_date_display = end_date_obj.strftime('%d/%m/%Y')
            except ValueError:
                return jsonify({
                    "success": False,
                    "error": "Định dạng ngày không hợp lệ. Sử dụng định dạng YYYY-MM-DD"
                }), 400

            # Generate mock report data
            report_data = []
            campaign_types = ['Ecommerce', 'Display Ads', 'Native Ads']
            ad_zones = ['Header Banner', 'Sidebar Right',
                        'In-feed Ads', 'Footer Banner', 'Pop-up Ads']

            for ad_id in ad_ids:
                # Generate daily data for the date range
                current_date = start_date_obj
                while current_date <= end_date_obj:
                    views = random.randint(100, 10000)
                    clicks = random.randint(10, views//10)
                    cost = random.randint(50000, 1000000)
                    purchases = random.randint(0, clicks//10)

                    daily_data = {
                        "ads_id": ad_id,
                        "ngay": current_date.strftime('%d/%m/%Y'),
                        "ten_tin_quang_cao": f"Quảng cáo mẫu {ad_id}",
                        "ten_chien_dich": f"Chiến dịch {(ad_id-1)//5 + 1}",
                        "ten_nhom": f"Nhóm quảng cáo {(ad_id-1)//3 + 1}",
                        "dinh_dang": random.choice(campaign_types),
                        "ten_vung_quang_cao": random.choice(ad_zones),
                        "luot_xem": views,
                        "luot_nhan": clicks,
                        "ctr": round((clicks / views * 100) if views > 0 else 0, 2),
                        "chi_phi": cost,
                        "cpc": round(cost/clicks if clicks > 0 else 0),
                        "cpm": round(cost*1000/views if views > 0 else 0),
                        "so_luong_mua_hang": purchases,
                        "ti_le_chuyen_doi": round((purchases / clicks * 100) if clicks > 0 else 0, 2),
                        "cps": round(cost/purchases if purchases > 0 else 0),
                        "video_view_3s": random.randint(10, views//2) if random.random() > 0.5 else 0
                    }
                    report_data.append(daily_data)
                    current_date += timedelta(days=1)

            # Generate Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Báo cáo quảng cáo"

            # Styles
            header_fill = PatternFill(
                start_color='0061C1', end_color='0061C1', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = [
                'Ngày', 'Tên tin quảng cáo', 'Chiến dịch', 'Nhóm', 'Định dạng',
                'Vùng quảng cáo', 'Lượt xem', 'Lượt nhấn', 'CTR (%)', 'Chi phí (VNĐ)',
                'CPC (VNĐ)', 'CPM (VNĐ)', 'Số lượng mua hàng', 'Tỷ lệ chuyển đổi (%)',
                'CPS (VNĐ)', 'Video view 3s'
            ]

            # Write title
            ws.merge_cells('A1:P1')
            title_cell = ws['A1']
            title_cell.value = f"Báo cáo quảng cáo từ {start_date_display} đến {end_date_display}"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Write data
            for row, data in enumerate(report_data, 3):
                ws.cell(row=row, column=1, value=data['ngay'])
                ws.cell(row=row, column=2, value=data['ten_tin_quang_cao'])
                ws.cell(row=row, column=3, value=data['ten_chien_dich'])
                ws.cell(row=row, column=4, value=data['ten_nhom'])
                ws.cell(row=row, column=5, value=data['dinh_dang'])
                ws.cell(row=row, column=6, value=data['ten_vung_quang_cao'])
                ws.cell(row=row, column=7, value=data['luot_xem'])
                ws.cell(row=row, column=8, value=data['luot_nhan'])
                ws.cell(row=row, column=9, value=data['ctr'])
                ws.cell(row=row, column=10, value=data['chi_phi'])
                ws.cell(row=row, column=11, value=data['cpc'])
                ws.cell(row=row, column=12, value=data['cpm'])
                ws.cell(row=row, column=13, value=data['so_luong_mua_hang'])
                ws.cell(row=row, column=14, value=data['ti_le_chuyen_doi'])
                ws.cell(row=row, column=15, value=data['cps'])
                ws.cell(row=row, column=16, value=data['video_view_3s'])

                # Apply borders to all cells in the row
                for col in range(1, 17):
                    ws.cell(row=row, column=col).border = border

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Generate filename
            filename = f"Bao_cao_tin_quang_cao_{start_date_display}_den_{end_date_display}.xlsx"
            filename = filename.replace('/', '_')

            # Create response
            response = make_response(excel_file.getvalue())
            response.headers.set(
                'Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response.headers.set(
                'Content-Disposition',
                f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
            )
            response.headers.set('Content-Length', len(excel_file.getvalue()))
            response.headers.set('Cache-Control', 'no-cache')
            response.headers.set('Pragma', 'no-cache')
            response.headers.set('Expires', '0')
            response.headers.set('X-Content-Type-Options', 'nosniff')

            return response

        except Exception as e:
            return jsonify({
                "success": False,
                "error": f"Lỗi hệ thống: {str(e)}"
            })
