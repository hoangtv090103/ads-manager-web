from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from configs.db import get_db_connection


class ZoneReport:
    @staticmethod
    def get_report(start_date=None, end_date=None, website_id=None, zone_id=None):
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()

                # Base query
                query = """
                    SELECT 
                        az.zone_id AS stt,
                        az.ten_vung_quang_cao AS ten_vung_quang_cao,
                        azs.ten_kich_thuoc AS kich_thuoc,
                        af.format_name AS dinh_dang,
                        CONCAT(w.domain_website, ' - ', p.email) AS website_publisher,
                        COALESCE(SUM(ad.luot_xem), 0) AS luot_xem,
                        COALESCE(SUM(ad.luot_nhan), 0) AS luot_nhan,
                        CAST(
                            CASE 
                                WHEN SUM(ad.luot_xem) > 0 
                                THEN ROUND((SUM(ad.luot_nhan)::float / SUM(ad.luot_xem) * 100)::numeric, 2)
                                ELSE 0 
                            END AS numeric
                        ) AS ctr,
                        COALESCE(SUM(ad.tong_chi_phi), 0) AS tong_gia_mua
                    FROM ads_zone az
                    LEFT JOIN website w ON az.website_id = w.website_id
                    LEFT JOIN publisher p ON w.publisher_id = p.publisher_id
                    LEFT JOIN ads_zone_size azs ON az.size_id = azs.size_id
                    LEFT JOIN ad_size_format asf ON azs.size_id = asf.size_id
                    LEFT JOIN ads_format af ON asf.format_id = af.format_id
                    LEFT JOIN ads_link_format alf ON af.format_id = alf.format_id
                    LEFT JOIN ads ad ON alf.ads_id = ad.ads_id
                    LEFT JOIN ads_group ag ON ad.ads_group_id = ag.ads_group_id
                    LEFT JOIN campaign c ON ag.camp_id = c.camp_id
                    WHERE az.active = TRUE 
                    AND w.active = TRUE
                    AND p.active = TRUE
                    AND (ad.active IS NULL OR ad.active = TRUE)
                    AND (ag.active IS NULL OR ag.active = TRUE)
                    AND (c.active IS NULL OR c.active = TRUE)
                """

                params = []

                # Add date filters
                if start_date:
                    query += " AND DATE(ad.created_at) >= %s"
                    params.append(start_date)
                if end_date:
                    query += " AND DATE(ad.created_at) <= %s"
                    params.append(end_date)

                # Add website filter
                if website_id:
                    query += " AND w.website_id = %s"
                    params.append(website_id)

                # Add zone filter
                if zone_id:
                    query += " AND az.zone_id = %s"
                    params.append(zone_id)

                # Group by and order
                query += """
                    GROUP BY 
                        az.zone_id,
                        az.ten_vung_quang_cao,
                        azs.ten_kich_thuoc,
                        af.format_name,
                        w.domain_website,
                        p.email
                    ORDER BY az.zone_id ASC
                """

                cursor.execute(query, params)
                results = cursor.fetchall()
                
                # Convert to dictionary
                columns = [desc[0] for desc in cursor.description]
                results = [dict(zip(columns, row)) for row in results]

                cursor.close()
                return results

        except Exception as e:
            raise Exception(f"Error getting zone report: {str(e)}")

    @staticmethod
    def generate_excel_report(report_data, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        
        # Company info
        ws['A1'] = 'Công ty cổ phần Novaon Digital'
        ws['A2'] = 'Địa chỉ: 94 Nguyễn Chính, Phương Liệt, Hoàng Mai, Hà Nội'
        ws['A3'] = 'Hotline: 024 011 5632'

        # Report title
        ws['A5'] = 'Báo cáo vùng quảng cáo'
        ws.merge_cells('A5:G5')
        ws['A5'].alignment = Alignment(horizontal='center')
        ws['A5'].font = Font(bold=True)

        # Report period
        period_text = f'Từ ngày {start_date} đến ngày {end_date}'
        ws['A6'] = period_text
        ws.merge_cells('A6:G6')
        ws['A6'].alignment = Alignment(horizontal='center')

        # Last updated info
        ws['A8'] = 'Số liệu được cập nhật lúc:'
        ws['B8'] = datetime.now().strftime('hh:mm:ss')
        ws['C8'] = datetime.now().strftime('dd/mm/yyyy')
        ws['A8'].font = Font(italic=True)

        # Headers
        headers = ['STT', 'Vùng quảng cáo', 'Website - Publisher', 'Lượt xem (qc)',
                  'Lượt nhấn (quảng cáo)', 'CTR (%)', 'Tổng giá mua (VNĐ)']
        header_row = 9
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='B8CCE4', 
                                  end_color='B8CCE4', 
                                  fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        if not report_data:
            # Nếu không có data, tạo 5 dòng trống
            for idx in range(1, 6):
                row = header_row + idx
                ws.cell(row=row, column=1, value=idx)
                for col in range(2, 8):
                    ws.cell(row=row, column=col, value='')
        else:
            for idx, data in enumerate(report_data, 1):
                row = header_row + idx
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=data['ten_vung_quang_cao'])
                ws.cell(row=row, column=3, value=f"{data['domain_website']} - {data['email']}")
                ws.cell(row=row, column=4, value=data['luot_xem'])
                ws.cell(row=row, column=5, value=data['luot_nhan'])
                ws.cell(row=row, column=6, value=data['ctr'])
                ws.cell(row=row, column=7, value=data['tong_gia_mua'])

        # Totals row
        total_row = header_row + (len(report_data) if report_data else 5) + 1
        ws.cell(row=total_row, column=1, value='Tổng:')
        ws.merge_cells(f'A{total_row}:C{total_row}')
        
        if report_data:
            total_views = sum(data['luot_xem'] for data in report_data)
            total_clicks = sum(data['luot_nhan'] for data in report_data)
            total_cost = sum(data['tong_gia_mua'] for data in report_data)
        else:
            total_views = 0
            total_clicks = 0
            total_cost = 0

        ws.cell(row=total_row, column=4, value=total_views)
        ws.cell(row=total_row, column=5, value=total_clicks)
        ws.cell(row=total_row, column=7, value=total_cost)

        # Style totals row
        for col in range(1, 8):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='B8CCE4', 
                                  end_color='B8CCE4',
                                  fill_type='solid')

        # Adjust column widths
        ws.column_dimensions['A'].width = 5   # STT
        ws.column_dimensions['B'].width = 30  # Zone Name  
        ws.column_dimensions['C'].width = 40  # Website - Publisher
        ws.column_dimensions['D'].width = 15  # Views
        ws.column_dimensions['E'].width = 15  # Clicks
        ws.column_dimensions['F'].width = 10  # CTR
        ws.column_dimensions['G'].width = 20  # Cost

        return wb
