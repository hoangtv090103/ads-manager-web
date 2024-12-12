from configs.db import get_db_connection
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

class AdFormatReport:
    def __init__(self, format_id=None, format_name=None, views=0, clicks=0, ctr=0, total_cost=0):
        self.format_id = format_id
        self.format_name = format_name
        self.views = views
        self.clicks = clicks
        self.ctr = ctr
        self.total_cost = total_cost

    @staticmethod
    def get_report(start_date=None, end_date=None, ad_formats=None):
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()

                # Base query
                query = """
                    SELECT
                        af.format_id AS id,
                        af.format_name AS format_name,
                        ct.ten_loai_chien_dich AS loai_chien_dich,
                        COALESCE(azs.ten_kich_thuoc, '-') AS kich_thuoc,
                        COALESCE(SUM(ad.tong_chi_phi), 0) AS chi_phi,
                        COALESCE(SUM(ad.luot_xem), 0) AS luot_xem,
                        COALESCE(SUM(ad.luot_nhan), 0) AS luot_nhan,
                        CAST(
                            CASE 
                                WHEN SUM(ad.luot_xem) > 0 
                                THEN ROUND((SUM(ad.luot_nhan)::float / SUM(ad.luot_xem) * 100)::numeric, 2)
                                ELSE 0 
                            END AS numeric
                        ) AS ctr,
                        CAST(
                            CASE 
                                WHEN SUM(ad.luot_nhan) > 0 
                                THEN ROUND((SUM(ad.tong_chi_phi)::float / SUM(ad.luot_nhan))::numeric, 2)
                                ELSE 0 
                            END AS numeric
                        ) AS cpc,
                        CAST(
                            CASE 
                                WHEN SUM(ad.luot_xem) > 0 
                                THEN ROUND((SUM(ad.tong_chi_phi)::float / SUM(ad.luot_xem) * 1000)::numeric, 2)
                                ELSE 0 
                            END AS numeric
                        ) AS cpm,
                        COALESCE(SUM(ad.so_luong_mua_hang), 0) AS sl_mua_hang,
                        CAST(
                            CASE 
                                WHEN SUM(ad.luot_nhan) > 0 
                                THEN ROUND((SUM(ad.so_luong_mua_hang)::float / SUM(ad.luot_nhan) * 100)::numeric, 2)
                                ELSE 0 
                            END AS numeric
                        ) AS ty_le_chuyen_doi,
                        COALESCE(SUM(ad.CPS), 0) AS cps
                    FROM ads_format af
                    LEFT JOIN campaign_type ct ON af.campaign_type_id = ct.camp_type_id
                    LEFT JOIN ad_size_format asf ON af.format_id = asf.format_id
                    LEFT JOIN ads_zone_size azs ON asf.size_id = azs.size_id
                    LEFT JOIN ads_link_format alf ON af.format_id = alf.format_id
                    LEFT JOIN ads ad ON alf.ads_id = ad.ads_id
                    LEFT JOIN ads_group ag ON ad.ads_group_id = ag.ads_group_id
                    WHERE af.active = TRUE
                    AND (ad.active IS NULL OR ad.active = TRUE)
                    AND (ag.active IS NULL OR ag.active = TRUE)
                """

                params = []

                # Add date filters
                if start_date:
                    query += " AND DATE(ad.created_at) >= %s"
                    params.append(start_date)
                if end_date:
                    query += " AND DATE(ad.created_at) <= %s"
                    params.append(end_date)

                # Add ad format filters
                if ad_formats:
                    query += " AND af.format_id IN %s"
                    params.append(tuple(ad_formats))

                query += """
                    GROUP BY 
                        af.format_id,
                        af.format_name,
                        ct.ten_loai_chien_dich,
                        azs.ten_kich_thuoc
                    ORDER BY af.format_id
                """

                cursor.execute(query, params)
                results = cursor.fetchall()
                
                # Convert to dictionary
                columns = [desc[0] for desc in cursor.description]
                results = [dict(zip(columns, row)) for row in results]

                cursor.close()
                return results

        except Exception as e:
            raise Exception(f"Error getting ad format report: {str(e)}")

    @staticmethod
    def generate_excel_report(report_data, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        
        # Company info
        ws['A1'] = 'Công ty cổ phần Novaon Digital'
        ws['A2'] = 'Địa chỉ: 94 Nguyễn Chính, Phương Liệt, Hoàng Mai, Hà Nội'
        ws['A3'] = 'Hotline: 024 011 5632'

        # Report title
        ws['A5'] = 'Báo cáo theo định dạng quảng cáo'
        ws.merge_cells('A5:M5')
        ws['A5'].alignment = Alignment(horizontal='center')
        ws['A5'].font = Font(bold=True)

        # Report period
        period_text = f'Từ ngày {start_date} đến ngày {end_date}'
        ws['A6'] = period_text
        ws.merge_cells('A6:M6')
        ws['A6'].alignment = Alignment(horizontal='center')

        # Headers
        headers = [
            'STT', 'Định dạng quảng cáo', 'Loại quảng cáo', 'Kích thước', 'Chi phí',
            'Lượt xem', 'Lượt nhấn', 'CTR', 'CPC', 'CPM', 'SL mua hàng',
            '% chuyển đổi mua hàng', 'CPS (VNĐ)'
        ]
        header_row = 7
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='B8CCE4', 
                                  end_color='B8CCE4', 
                                  fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        if not report_data:
            # Nếu không có data, tạo 5 dòng trống
            for idx in range(1, 6):
                row = header_row + idx
                ws.cell(row=row, column=1, value=idx)
                for col in range(2, 14):
                    ws.cell(row=row, column=col, value='')
        else:
            for idx, data in enumerate(report_data, 1):
                row = header_row + idx
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=data['format_name'])
                ws.cell(row=row, column=3, value=data['loai_chien_dich'])
                ws.cell(row=row, column=4, value=data['kich_thuoc'])
                ws.cell(row=row, column=5, value=data['chi_phi'])
                ws.cell(row=row, column=6, value=data['luot_xem'])
                ws.cell(row=row, column=7, value=data['luot_nhan'])
                ws.cell(row=row, column=8, value=data['ctr'])
                ws.cell(row=row, column=9, value=data['cpc'])
                ws.cell(row=row, column=10, value=data['cpm'])
                ws.cell(row=row, column=11, value=data['sl_mua_hang'])
                ws.cell(row=row, column=12, value=data['ty_le_chuyen_doi'])
                ws.cell(row=row, column=13, value=data['cps'])

        # Totals row
        total_row = header_row + (len(report_data) if report_data else 5) + 1
        ws.cell(row=total_row, column=1, value='Tổng:')
        ws.cell(row=total_row, column=2, value=len(report_data))

        # Style totals row
        for col in range(1, 14):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='B8CCE4', 
                                  end_color='B8CCE4',
                                  fill_type='solid')

        # Adjust column widths
        ws.column_dimensions['A'].width = 5   # STT
        ws.column_dimensions['B'].width = 25  # Định dạng quảng cáo
        ws.column_dimensions['C'].width = 20  # Loại quảng cáo
        ws.column_dimensions['D'].width = 15  # Kích thước
        ws.column_dimensions['E'].width = 15  # Chi phí
        ws.column_dimensions['F'].width = 12  # Lượt xem
        ws.column_dimensions['G'].width = 12  # Lượt nhấn
        ws.column_dimensions['H'].width = 10  # CTR
        ws.column_dimensions['I'].width = 10  # CPC
        ws.column_dimensions['J'].width = 10  # CPM
        ws.column_dimensions['K'].width = 15  # SL mua hàng
        ws.column_dimensions['L'].width = 20  # % chuyển đổi
        ws.column_dimensions['M'].width = 15  # CPS

        return wb