from configs.db import get_db_connection
from openpyxl import load_workbook
import io


class Product:
    def __init__(self, product_id=None, ten_san_pham="", mo_ta_san_pham="", lien_ket_san_pham="", hinh_anh_san_pham=None, productstatus_id="", tinh_trang_san_pham="", gia=0.0, gia_khuyen_mai=0.0, loai_san_pham="", customer_id=""):
        self.product_id = product_id
        self.ten_san_pham = ten_san_pham
        self.mo_ta_san_pham = mo_ta_san_pham
        self.lien_ket_san_pham = lien_ket_san_pham
        self.hinh_anh_san_pham = hinh_anh_san_pham
        self.productstatus_id = productstatus_id
        self.tinh_trang_san_pham = tinh_trang_san_pham
        self.gia = gia
        self.gia_khuyen_mai = gia_khuyen_mai
        self.loai_san_pham = loai_san_pham
        self.customer_id = customer_id

    @staticmethod
    def create_table():
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS product (
                    product_id SERIAL PRIMARY KEY,
                    ten_san_pham VARCHAR(50),
                    mo_ta_san_pham VARCHAR(50),
                    lien_ket_san_pham VARCHAR(50),
                    hinh_anh_san_pham BYTEA,
                    productstatus_id INTEGER,
                    tinh_trang_san_pham VARCHAR(50),
                    gia FLOAT,
                    gia_khuyen_mai FLOAT,
                    loai_san_pham VARCHAR(50),
                    customer_id INTEGER,
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    active BOOLEAN DEFAULT TRUE,
                    FOREIGN KEY (productstatus_id) REFERENCES product_status(status_id)
                        ON DELETE CASCADE,
                    FOREIGN KEY (customer_id) REFERENCES customer(customer_id)
                        ON DELETE CASCADE
                )
            ''')
            conn.commit()

    def save(self):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
            INSERT INTO product (product_id, ten_san_pham, mo_ta_san_pham, lien_ket_san_pham, hinh_anh_san_pham, productstatus_id, tinh_trang_san_pham, gia, gia_khuyen_mai, loai_san_pham, customer_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (self.product_id, self.ten_san_pham, self.mo_ta_san_pham, self.lien_ket_san_pham, self.hinh_anh_san_pham, self.productstatus_id, self.tinh_trang_san_pham, self.gia, self.gia_khuyen_mai, self.loai_san_pham, self.customer_id))
            conn.commit()

    @staticmethod
    def get_all():
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT p.*, 
                       COALESCE(s.luot_xem, 0) as luot_xem,
                       COALESCE(s.luot_nhan, 0) as luot_nhan,
                       CASE 
                           WHEN s.luot_xem > 0 
                           THEN ROUND(CAST(s.luot_nhan AS FLOAT) / s.luot_xem * 100, 2)
                           ELSE 0 
                       END as ctr
                FROM product p
                LEFT JOIN product_stats s ON p.product_id = s.product_id
                WHERE p.active = true
            ''')
            rows = cursor.fetchall()
            return [dict(zip([column[0] for column in cursor.description], row)) for row in rows]

    def update(self):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
            UPDATE product
            SET ten_san_pham = ?, mo_ta_san_pham = ?, lien_ket_san_pham = ?, hinh_anh_san_pham = ?, productstatus_id = ?, tinh_trang_san_pham = ?, gia = ?, gia_khuyen_mai = ?, loai_san_pham = ?, customer_id = ?
            WHERE product_id = ?
            ''', (self.ten_san_pham, self.mo_ta_san_pham, self.lien_ket_san_pham, self.hinh_anh_san_pham, self.productstatus_id, self.tinh_trang_san_pham, self.gia, self.gia_khuyen_mai, self.loai_san_pham, self.customer_id, self.product_id))
            conn.commit()

    @staticmethod
    def delete_by_id(product_id):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
            UPDATE product SET active = 0 WHERE product_id = ?
            ''', (product_id,))
            conn.commit()

    @staticmethod
    def import_from_excel(file_content):
        wb = load_workbook(filename=io.BytesIO(file_content))
        ws = wb.active
        products = []

        for row in ws.iter_rows(min_row=2):  # Skip header row
            product = Product(
                ten_san_pham=row[0].value,
                mo_ta_san_pham=row[1].value,
                lien_ket_san_pham=row[2].value,
                gia=float(row[3].value or 0),
                gia_khuyen_mai=float(row[4].value or 0),
                loai_san_pham=row[5].value
            )
            products.append(product)
            product.save()

        return products
