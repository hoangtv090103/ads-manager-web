from configs.db import get_db_connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
from collections import defaultdict
import calendar


class AdsReport:
    @staticmethod
    def get_report_data(ad_ids, start_date=None, end_date=None):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            query = '''
                WITH ads_metrics AS (
                    SELECT
                        a.ads_id,
                        a.ten_tin_quang_cao as "Tên tin QC",
                        c.ten_chien_dich as "Chiến dịch", 
                        ag.ten_nhom as "Tên nhóm QC",
                        af.format_name as "Loại quảng cáo",
                        azs.ten_kich_thuoc as "Kích thước quảng cáo",
                        a.url_dich as "URL Đích",
                        a.tong_chi_phi as "Chi phí(VNĐ)",
                        a.luot_xem as "Lượt xem",
                        a.luot_nhan as "Lượt nhấn",
                        CASE 
                            WHEN a.luot_nhan > 0 THEN CAST((a.video_watches_at_3s::float / a.luot_nhan * 100) AS DECIMAL(10,2))
                            ELSE 0 
                        END as "% Video view 3s",
                        CASE 
                            WHEN a.luot_xem > 0 THEN CAST((a.luot_nhan::float / a.luot_xem * 100) AS DECIMAL(10,2))
                            ELSE 0 
                        END as "CTR(%)",
                        CASE 
                            WHEN a.luot_nhan > 0 THEN CAST((a.tong_chi_phi::float / a.luot_nhan) AS DECIMAL(10,2))
                            ELSE 0 
                        END as "CPC (VNĐ)",
                        CASE 
                            WHEN a.luot_xem > 0 THEN CAST((a.tong_chi_phi::float / a.luot_xem * 1000) AS DECIMAL(10,2))
                            ELSE 0 
                        END as "CPM (VNĐ)",
                        a.so_luong_mua_hang as "SL mua hàng",
                        CASE 
                            WHEN a.luot_nhan > 0 THEN CAST((a.so_luong_mua_hang::float / a.luot_nhan * 100) AS DECIMAL(10,2))
                            ELSE 0 
                        END as "% chuyển đổi mua hàng",
                        CASE 
                            WHEN a.so_luong_mua_hang > 0 THEN CAST((a.tong_chi_phi::float / a.so_luong_mua_hang) AS DECIMAL(10,2))
                            ELSE 0 
                        END as "CPS (VNĐ)",
                        a.video_watches_at_3s as "Video view 3s",
                        a.video_watches_at_25 as "Video watches at 25%",
                        a.video_watches_at_50 as "Video watches at 50%", 
                        a.video_watches_at_75 as "Video watches at 75%",
                        a.video_watches_at_100 as "Video watches at 100%",
                        ast.ten_trang_thai as "Trạng thái",
                        c.ngay_bat_dau as "Ngày bắt đầu chiến dịch",
                        c.ngay_ket_thuc as "Ngày kết thúc chiến dịch",
                        EXTRACT(MONTH FROM c.ngay_bat_dau) as month,
                        EXTRACT(YEAR FROM c.ngay_bat_dau) as year,
                        EXTRACT(WEEK FROM c.ngay_bat_dau) as week,
                        DATE(c.ngay_bat_dau) as date
                    FROM ads a
                    LEFT JOIN ads_group ag ON ag.ads_group_id = a.ads_group_id
                    LEFT JOIN campaign c ON c.camp_id = ag.camp_id
                    LEFT JOIN status_ads ast ON ast.status_id = a.ads_status_id
                    LEFT JOIN ads_link_format alf ON alf.ads_id = a.ads_id
                    LEFT JOIN ads_format af ON af.format_id = alf.format_id
                    LEFT JOIN ad_size_format asf ON asf.format_id = alf.format_id
                    LEFT JOIN ads_zone_size azs ON azs.size_id = asf.size_id
                    WHERE a.ads_id = ANY(%s::integer[])
                    AND a.active = true
                    AND ag.active = true
                )
                SELECT * FROM ads_metrics
            '''
            params = [ad_ids]

            if start_date and end_date:
                query += ' WHERE date >= %s AND date <= %s'
                params.extend([start_date, end_date])

            query += ' ORDER BY domain_website, date'

            cursor.execute(query, params)
            return [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]

    @staticmethod
    def generate_excel_report(report_data, start_date, end_date):
        wb = Workbook()

        # Tạo các sheets
        overview_sheet = wb.active
        overview_sheet.title = "Bao cao tong quan tin quang cao"
        monthly_sheet = wb.create_sheet(
            "Bao cao chi tiet tin quang cao theo thang")
        weekly_sheet = wb.create_sheet("Bao cao tin quang cao theo tuan")
        daily_sheet = wb.create_sheet("Bao cao tin quang cao theo ngay")
        website_sheet = wb.create_sheet("Bao cao tin quang cao theo website")

        # Định nghĩa styles
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def add_common_header(sheet):
            sheet['A1'] = 'Cong ty co phan Novaon Digital'
            sheet['A2'] = 'Dia chi: 94 Nguyen Chinh, Phuong Liet, Hoang Mai, Ha Noi'
            sheet['A3'] = 'Hotline: 024 011 5632'

            sheet['A6'] = f'Tu ngay {start_date} den ngay {end_date}'
            sheet.merge_cells('A6:G6')
            sheet['A6'].alignment = center_align

            for row in range(1, 4):
                cell = sheet[f'A{row}']
                cell.font = Font(bold=True if row == 1 else False)

        def setup_monthly_sheet():
            add_common_header(monthly_sheet)
            monthly_sheet['A5'] = 'Bao cao chi tiet tin quang cao theo thang'
            monthly_sheet.merge_cells('A5:G5')
            monthly_sheet['A5'].alignment = center_align
            monthly_sheet['A5'].font = header_font

            # Headers cho bảng
            headers = ['Thang', 'Ten tin quang cao', 'Ten chien dich', 'Ten nhom', 'Dinh dang', 'URL dich',
                       'Trang thai', 'Chi phi (VND)', 'Luot xem', 'Luot nhan', 'CTR(%)', 'CPC (VND)',
                       'CPM (VND)', 'SL mua hang', '% chuyen doi mua hang', 'CPS (VND)', 'Video view 3s',
                       'Video watches at 25%', 'Video watches at 50%', 'Video watches at 75%',
                       'Video watches at 100%', 'Trang thai hien thi', 'Vung quang cao']

            # Organize data by month
            monthly_data = defaultdict(list)
            for item in report_data:
                month_key = f"{int(item['month'])}/{int(item['year'])}"
                monthly_data[month_key].append(item)

            current_row = 10
            for month_key in sorted(monthly_data.keys(), key=lambda x: datetime.strptime(x, "%m/%Y")):
                # Add month header
                for col, header in enumerate(headers, 1):
                    cell = monthly_sheet.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                current_row += 1
                month_data = monthly_data[month_key]

                # Add data for each ad in this month
                for data in month_data:
                    monthly_sheet.cell(
                        row=current_row, column=1, value=month_key)
                    monthly_sheet.cell(
                        row=current_row, column=2, value=data['ten_tin_quang_cao'])
                    monthly_sheet.cell(
                        row=current_row, column=3, value=data['ten_chien_dich'])
                    monthly_sheet.cell(
                        row=current_row, column=4, value=data['ten_nhom'])
                    monthly_sheet.cell(
                        row=current_row, column=5, value=data['dinh_dang'])
                    monthly_sheet.cell(
                        row=current_row, column=6, value=data['url_dich'])
                    monthly_sheet.cell(
                        row=current_row, column=7, value=data['trang_thai'])
                    monthly_sheet.cell(
                        row=current_row, column=8, value=data['tong_chi_phi'])
                    monthly_sheet.cell(
                        row=current_row, column=9, value=data['luot_xem'])
                    monthly_sheet.cell(
                        row=current_row, column=10, value=data['luot_nhan'])
                    monthly_sheet.cell(
                        row=current_row, column=11, value=data['ctr'])
                    monthly_sheet.cell(
                        row=current_row, column=12, value=data['cpc'])
                    monthly_sheet.cell(
                        row=current_row, column=13, value=data['cpm'])
                    monthly_sheet.cell(
                        row=current_row, column=14, value=data['so_luong_mua_hang'])
                    monthly_sheet.cell(
                        row=current_row, column=15, value=data['conversion_rate'])
                    monthly_sheet.cell(
                        row=current_row, column=16, value=data['cps'])
                    monthly_sheet.cell(
                        row=current_row, column=17, value=data['videoview3s'])
                    monthly_sheet.cell(
                        row=current_row, column=18, value=data['videowatchesat25'])
                    monthly_sheet.cell(
                        row=current_row, column=19, value=data['videowatchesat50'])
                    monthly_sheet.cell(
                        row=current_row, column=20, value=data['videowatchesat75'])
                    monthly_sheet.cell(
                        row=current_row, column=21, value=data['videowatchesat100'])
                    monthly_sheet.cell(
                        row=current_row, column=22, value='Đang hiển thị' if data['dang_hien_thi'] else 'Chưa hiển thị')
                    monthly_sheet.cell(
                        row=current_row, column=23, value=data['ten_vung_quang_cao'])

                    # Apply borders and alignment
                    for col in range(1, 24):
                        cell = monthly_sheet.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')

                    current_row += 1

        def setup_weekly_sheet():
            add_common_header(weekly_sheet)
            weekly_sheet['A5'] = 'Bao cao tin quang cao theo tuan'
            weekly_sheet.merge_cells('A5:G5')
            weekly_sheet['A5'].alignment = center_align
            weekly_sheet['A5'].font = header_font

            # Headers cho bảng
            headers = ['Tuan', 'Ten tin quang cao', 'Ten chien dich', 'Ten nhom', 'Dinh dang', 'URL dich',
                       'Trang thai', 'Chi phi (VND)', 'Luot xem', 'Luot nhan', 'CTR(%)', 'CPC (VND)',
                       'CPM (VND)', 'SL mua hang', '% chuyen doi mua hang', 'CPS (VND)', 'Video view 3s',
                       'Video watches at 25%', 'Video watches at 50%', 'Video watches at 75%',
                       'Video watches at 100%', 'Trang thai hien thi', 'Vung quang cao']

            # Similar structure as monthly sheet but organized by week
            # Implementation similar to monthly sheet but with week-based grouping

        def setup_daily_sheet():
            add_common_header(daily_sheet)
            daily_sheet['A5'] = 'Bao cao tin quang cao theo ngay'
            daily_sheet.merge_cells('A5:G5')
            daily_sheet['A5'].alignment = center_align
            daily_sheet['A5'].font = header_font

            # Headers cho bảng
            headers = ['Ngay', 'Ten tin quang cao', 'Ten chien dich', 'Ten nhom', 'Dinh dang', 'URL dich',
                       'Trang thai', 'Chi phi (VND)', 'Luot xem', 'Luot nhan', 'CTR(%)', 'CPC (VND)',
                       'CPM (VND)', 'SL mua hang', '% chuyen doi mua hang', 'CPS (VND)', 'Video view 3s',
                       'Video watches at 25%', 'Video watches at 50%', 'Video watches at 75%',
                       'Video watches at 100%', 'Trang thai hien thi', 'Vung quang cao']

            # Similar structure as monthly sheet but organized by day
            # Implementation similar to monthly sheet but with day-based grouping

        def setup_website_sheet():
            add_common_header(website_sheet)
            website_sheet['A5'] = 'Bao cao tin quang cao theo website'
            website_sheet.merge_cells('A5:G5')
            website_sheet['A5'].alignment = center_align
            website_sheet['A5'].font = header_font

            # Headers cho bảng
            headers = ['Website', 'Ten tin quang cao', 'Ten chien dich', 'Ten nhom', 'Dinh dang', 'URL dich',
                       'Trang thai', 'Chi phi (VND)', 'Luot xem', 'Luot nhan', 'CTR(%)', 'CPC (VND)',
                       'CPM (VND)', 'SL mua hang', '% chuyen doi mua hang', 'CPS (VND)', 'Video view 3s',
                       'Video watches at 25%', 'Video watches at 50%', 'Video watches at 75%',
                       'Video watches at 100%', 'Trang thai hien thi', 'Vung quang cao']

            # Similar structure as monthly sheet but organized by website
            # Implementation similar to monthly sheet but with website-based grouping

        # Thực hiện setup các sheets
        setup_monthly_sheet()
        setup_weekly_sheet()
        setup_daily_sheet()
        setup_website_sheet()

        # Điều chỉnh độ rộng cột cho tất cả các sheets
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

        return wb
