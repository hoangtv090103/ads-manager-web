from configs.db import get_db_connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class WebsiteReport:
    def __init__(self, website_id=None, publisher_email=None, views=0, clicks=0, ctr=0, total_cost=0):
        self.website_id = website_id
        self.publisher_email = publisher_email
        self.views = views
        self.clicks = clicks
        self.ctr = ctr
        self.total_cost = total_cost

    @staticmethod
    def get_report(start_date=None, end_date=None, publisher_email=None):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            query = '''
                SELECT 
                    w.website_id,
                    w.domain_website,
                    p.email as publisher_email,
                    SUM(c.luot_xem) as total_views,
                    SUM(c.luot_nhan) as total_clicks,
                    CASE 
                        WHEN SUM(c.luot_xem) > 0 
                        THEN ROUND(((SUM(c.luot_nhan)::numeric / SUM(c.luot_xem)) * 100), 2)
                        ELSE 0 
                    END as ctr,
                    SUM(c.tong_chi_phi) as total_cost
                FROM website w
                JOIN publisher p ON w.publisher_id = p.publisher_id
                JOIN ads_group_website agw ON w.website_id = agw.website_id
                JOIN ads_group ag ON agw.ads_group_id = ag.ads_group_id
                JOIN campaign c ON ag.camp_id = c.camp_id
                WHERE c.active = true
                    AND w.active = true
                    AND p.active = true
                    AND agw.active = true
                    AND ag.active = true
            '''
            params = []

            if start_date and end_date:
                query += ' AND c.ngay_bat_dau >= %s AND c.ngay_ket_thuc <= %s'
                params.extend([start_date, end_date])

            if publisher_email:
                query += ' AND p.email = %s'
                params.append(publisher_email)

            query += '''
                GROUP BY w.website_id, w.domain_website, p.email
                ORDER BY w.website_id
            '''

            cursor.execute(query, params)
            rows = cursor.fetchall()
            return [dict(zip([column[0] for column in cursor.description], row)) for row in rows]

    @staticmethod
    def generate_excel_report(report_data, start_date, end_date):
        wb = Workbook()
        ws = wb.active

        # Company info
        ws['A1'] = 'Công ty cổ phần Novaon Digital'
        ws['A2'] = 'Địa chỉ: 94 Nguyễn Chính, Phương Liệt, Hoàng Mai, Hà Nội'
        ws['A3'] = 'Hotline: 024 011 5632'

        # Report title
        ws['A4'] = 'Báo cáo Publisher'
        ws['A4'].font = Font(size=16, bold=True)
        ws.merge_cells('A4:G4')
        ws['A4'].alignment = Alignment(horizontal='center')

        # Report period
        period_text = f'Từ ngày {start_date} đến ngày {end_date}'
        ws['A5'] = period_text
        ws.merge_cells('A5:G5')
        ws['A5'].alignment = Alignment(horizontal='center')

        # Last updated info
        ws['A7'] = 'Số liệu được cập nhật lúc:'
        ws['B7'] = datetime.now().strftime('%H:%M:%S')
        ws['C7'] = datetime.now().strftime('%d/%m/%Y')
        ws['A7'].font = Font(italic=True)

        # Headers
        headers = ['STT', 'Tên publisher', 'Email', 'Lượt xem',
                   'Lượt nhấn (quảng cáo)', 'CTR (%)', 'Tổng giá mua (VNĐ)']
        header_row = 8
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        if not report_data:
            # Nếu không có data, tạo 5 dòng trống
            for idx in range(1, 6):
                row = header_row + idx
                ws.cell(row=row, column=1, value=idx)
                for col in range(2, 8):
                    ws.cell(row=row, column=col, value='')
        else:
            for idx, data in enumerate(report_data, 1):
                row = header_row + idx
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=data['domain_website'])
                ws.cell(row=row, column=3, value=data['publisher_email'])
                ws.cell(row=row, column=4, value=data['total_views'])
                ws.cell(row=row, column=5, value=data['total_clicks'])
                ws.cell(row=row, column=6, value=data['ctr'])
                ws.cell(row=row, column=7, value=data['total_cost'])

        # Totals row
        total_row = header_row + (len(report_data) if report_data else 5) + 1
        ws.cell(row=total_row, column=1, value='Tổng:')
        ws.merge_cells(f'A{total_row}:B{total_row}')
        
        if report_data:
            total_records = len(report_data)
            total_views = sum(data['total_views'] for data in report_data)
            total_clicks = sum(data['total_clicks'] for data in report_data)
            total_cost = sum(data['total_cost'] for data in report_data)
        else:
            total_records = 0
            total_views = 0
            total_clicks = 0
            total_cost = 0

        ws.cell(row=total_row, column=3, value=total_records)
        ws.cell(row=total_row, column=4, value=total_views)
        ws.cell(row=total_row, column=5, value=total_clicks)
        ws.cell(row=total_row, column=7, value=total_cost)

        # Style totals row
        for col in range(1, 8):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        return wb
