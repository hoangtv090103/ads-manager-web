from configs.db import get_db_connection
from openpyxl import load_workbook
import io
import base64


class Product:
    def __init__(self, product_id=None, ten_san_pham="", mo_ta_san_pham="",
                 luot_xem=None, luot_nhan=None, ctr=None,
                 so_luong_mua_hang=None, phan_tram_chuyen_doi_mua_hang=None,
                 lien_ket_san_pham="", hinh_anh_san_pham=None,
                 lien_ket_hinh_anh="", source_id=None, productstatus_id=None,
                 gia_san_pham=None, gia_khuyen_mai=None,
                 created_at=None, updated_at=None, active=True):
        self.product_id = product_id
        self.ten_san_pham = ten_san_pham
        self.mo_ta_san_pham = mo_ta_san_pham
        self.luot_xem = luot_xem
        self.luot_nhan = luot_nhan
        self.ctr = ctr
        self.so_luong_mua_hang = so_luong_mua_hang
        self.phan_tram_chuyen_doi_mua_hang = phan_tram_chuyen_doi_mua_hang
        self.lien_ket_san_pham = lien_ket_san_pham
        self.hinh_anh_san_pham = hinh_anh_san_pham
        self.lien_ket_hinh_anh = lien_ket_hinh_anh
        self.source_id = source_id
        self.productstatus_id = productstatus_id
        self.gia_san_pham = gia_san_pham
        self.gia_khuyen_mai = gia_khuyen_mai
        self.created_at = created_at
        self.updated_at = updated_at
        self.active = active

    @staticmethod
    def create_table():
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS product (
                    product_id SERIAL PRIMARY KEY,
                    ten_san_pham VARCHAR(255),
                    mo_ta_san_pham TEXT,
                    luot_xem INTEGER,
                    luot_nhan INTEGER,
                    ctr DECIMAL(15,2),
                    so_luong_mua_hang INTEGER,
                    phan_tram_chuyen_doi_mua_hang DECIMAL(15,2),
                    lien_ket_san_pham TEXT,
                    hinh_anh_san_pham BYTEA,
                    lien_ket_hinh_anh TEXT,
                    source_id INTEGER,
                    productstatus_id INTEGER,
                    gia_san_pham DECIMAL(15,2),
                    gia_khuyen_mai DECIMAL(15,2),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    active BOOLEAN DEFAULT TRUE,
                    FOREIGN KEY (source_id) REFERENCES data_source(source_id)
                        ON DELETE CASCADE,
                    FOREIGN KEY (productstatus_id) REFERENCES product_status(productstatus_id)
                        ON DELETE CASCADE
                )
            ''')
            conn.commit()

    def save(self):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
            INSERT INTO product (
                ten_san_pham, mo_ta_san_pham, luot_xem, luot_nhan, ctr,
                so_luong_mua_hang, phan_tram_chuyen_doi_mua_hang, lien_ket_san_pham,
                hinh_anh_san_pham, lien_ket_hinh_anh, source_id, productstatus_id,
                gia_san_pham, gia_khuyen_mai, active
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING product_id
            ''', (
                self.ten_san_pham, self.mo_ta_san_pham,
                self.luot_xem, self.luot_nhan, self.ctr,
                self.so_luong_mua_hang, self.phan_tram_chuyen_doi_mua_hang,
                self.lien_ket_san_pham, self.hinh_anh_san_pham,
                self.lien_ket_hinh_anh, self.source_id, self.productstatus_id, 
                self.gia_san_pham, self.gia_khuyen_mai, self.active
            ))
            product_id = cursor.fetchone()[0]
            conn.commit()
            return product_id

    @staticmethod
    def get_all():
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT p.*, 
                           ds.ten_nguon_du_lieu,
                           ps.ten_trang_thai as tinh_trang_san_pham
                    FROM product p
                    LEFT JOIN data_source ds ON p.source_id = ds.source_id
                    LEFT JOIN product_status ps ON p.productstatus_id = ps.productstatus_id
                    WHERE p.active = true
                    ORDER BY p.created_at DESC
                ''')
                rows = cursor.fetchall()
                if not rows:
                    return []
                
                result = []
                for row in rows:
                    product_dict = dict(zip([column[0] for column in cursor.description], row))
                    # Format numeric values
                    product_dict['gia_san_pham'] = float(product_dict['gia_san_pham']) if product_dict['gia_san_pham'] else 0
                    product_dict['gia_khuyen_mai'] = float(product_dict['gia_khuyen_mai']) if product_dict['gia_khuyen_mai'] else None
                    product_dict['ctr'] = float(product_dict['ctr']) if product_dict['ctr'] else 0
                    # Convert bytea to base64 if exists
                    if product_dict['hinh_anh_san_pham']:
                        product_dict['hinh_anh_san_pham'] = base64.b64encode(product_dict['hinh_anh_san_pham']).decode('utf-8')
                    # Add default value for so_nhom_quang_cao since we removed the JOIN
                    product_dict['so_nhom_quang_cao'] = 0
                    result.append(product_dict)
                return result
        except Exception as e:
            print(f"Error getting all products: {str(e)}")
            raise

    @staticmethod
    def get_by_id(product_id):
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT p.*, 
                           ds.ten_nguon_du_lieu,
                           ps.ten_trang_thai as tinh_trang_san_pham
                    FROM product p
                    LEFT JOIN data_source ds ON p.source_id = ds.source_id
                    LEFT JOIN product_status ps ON p.productstatus_id = ps.productstatus_id
                    WHERE p.product_id = %s
                ''', (product_id,))
                row = cursor.fetchone()
                if not row:
                    return None
                
                product_dict = dict(zip([column[0] for column in cursor.description], row))
                # Format numeric values
                product_dict['gia_san_pham'] = float(product_dict['gia_san_pham']) if product_dict['gia_san_pham'] else 0
                product_dict['gia_khuyen_mai'] = float(product_dict['gia_khuyen_mai']) if product_dict['gia_khuyen_mai'] else None
                product_dict['ctr'] = float(product_dict['ctr']) if product_dict['ctr'] else 0
                # Convert bytea to base64 if exists
                if product_dict['hinh_anh_san_pham']:
                    product_dict['hinh_anh_san_pham'] = base64.b64encode(product_dict['hinh_anh_san_pham']).decode('utf-8')
                # Add default value for so_nhom_quang_cao since we removed the JOIN
                product_dict['so_nhom_quang_cao'] = 0
                return product_dict
        except Exception as e:
            print(f"Error getting product by ID: {str(e)}")
            raise

    @staticmethod
    def update_status(product_id, active):
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE product 
                    SET active = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE product_id = %s
                ''', (active, product_id))
                conn.commit()
        except Exception as e:
            print(f"Error updating product status: {str(e)}")
            raise

    def update(self):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
            UPDATE product
            SET ten_san_pham = %s,
                mo_ta_san_pham = %s,
                luot_xem = %s,
                luot_nhan = %s,
                ctr = %s,
                so_luong_mua_hang = %s,
                phan_tram_chuyen_doi_mua_hang = %s,
                lien_ket_san_pham = %s,
                hinh_anh_san_pham = %s,
                source_id = %s,
                productstatus_id = %s,
                gia_san_pham = %s,
                gia_khuyen_mai = %s,
                bo_chi_so = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE product_id = %s
            ''', (
                self.ten_san_pham, self.mo_ta_san_pham,
                self.luot_xem, self.luot_nhan, self.ctr,
                self.so_luong_mua_hang, self.phan_tram_chuyen_doi_mua_hang,
                self.lien_ket_san_pham, self.hinh_anh_san_pham,
                self.source_id, self.productstatus_id,
                self.gia_san_pham, self.gia_khuyen_mai,
                self.product_id
            ))
            conn.commit()

    @staticmethod
    def delete_by_id(product_id):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
            UPDATE product 
            SET active = false,
                updated_at = CURRENT_TIMESTAMP
            WHERE product_id = %s
            ''', (product_id,))
            conn.commit()

    @staticmethod
    def import_from_excel(file_content):
        wb = load_workbook(filename=io.BytesIO(file_content))
        ws = wb.active
        products = []

        for row in ws.iter_rows(min_row=2):  # Skip header row
            product = Product(
                ten_san_pham=row[0].value,
                mo_ta_san_pham=row[1].value,
                bo_chi_so=row[2].value,
                lien_ket_san_pham=row[3].value,
                hinh_anh_san_pham=row[4].value,
                lien_ket_hinh_anh=row[5].value,
                source_id=row[6].value,
                productstatus_id=row[7].value,
                gia_san_pham=row[8].value,
                gia_khuyen_mai=row[9].value
            )
            products.append(product)
            product.save()

        return products
