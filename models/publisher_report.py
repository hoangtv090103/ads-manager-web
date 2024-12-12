from configs.db import get_db_connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

class PublisherReport:
    @staticmethod
    def get_report(start_date=None, end_date=None, publisher_email=None):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            query = '''
                SELECT 
                    p.publisher_id,
                    p.ten_publisher,
                    p.email as publisher_email,
                    SUM(c.luot_xem) as total_views,
                    SUM(c.luot_nhan) as total_clicks,
                    CASE 
                        WHEN SUM(c.luot_xem) > 0 
                        THEN ROUND(((SUM(c.luot_nhan)::numeric / SUM(c.luot_xem)) * 100), 2)
                        ELSE 0 
                    END as ctr,
                    SUM(c.tong_chi_phi) as total_cost
                FROM publisher p
                JOIN campaign c ON c.publisher_id = p.publisher_id 
                WHERE c.active = true
                    AND p.active = true
            '''
            params = []
            
            if start_date and end_date:
                query += ' AND c.ngay_bat_dau >= %s AND c.ngay_ket_thuc <= %s'
                params.extend([start_date, end_date])
            
            if publisher_email:
                query += ' AND p.email = %s'
                params.append(publisher_email)
                
            query += '''
                GROUP BY p.publisher_id, p.ten_publisher, p.email
                ORDER BY p.publisher_id
            '''
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            return [dict(zip([column[0] for column in cursor.description], row)) for row in rows]

    @staticmethod
    def generate_excel_report(report_data, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        
        # Company info
        ws['A1'] = 'Công ty cổ phần Novaon Digital'
        ws['A2'] = 'Địa chỉ: 94 Nguyễn Chính, Phương Liệt, Hoàng Mai, Hà Nội'
        ws['A3'] = 'Hotline: 024 011 5632'
        
        # Report title
        ws['A5'] = 'Báo cáo Publisher'
        ws.merge_cells('A5:G5')
        ws['A5'].alignment = Alignment(horizontal='center')
        ws['A5'].font = Font(bold=True)
        
        # Report period
        period_text = f'Từ ngày {start_date} đến ngày {end_date}'
        ws['A6'] = period_text
        ws.merge_cells('A6:G6')
        ws['A6'].alignment = Alignment(horizontal='center')
        
        # Last updated info
        ws['A8'] = 'Số liệu được cập nhật lúc:'
        ws['B8'] = datetime.now().strftime('hh:mm:ss')
        ws['C8'] = datetime.now().strftime('dd/mm/yyyy')
        ws['A8'].font = Font(italic=True)
        
        # Headers
        headers = ['STT', 'Tên publisher', 'Email', 'Lượt xem',
                  'Lượt nhấn (quảng cáo)', 'CTR (%)', 'Tổng giá mua (VNĐ)']
        header_row = 9
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        if not report_data:
            # Nếu không có data, tạo 5 dòng trống
            for idx in range(1, 6):
                row = header_row + idx
                ws.cell(row=row, column=1, value=idx)
                for col in range(2, 8):
                    ws.cell(row=row, column=col, value='')
        else:
            for idx, data in enumerate(report_data, 1):
                row = header_row + idx
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=data['ten_publisher'])
                ws.cell(row=row, column=3, value=data['publisher_email'])
                ws.cell(row=row, column=4, value=data['total_views'])
                ws.cell(row=row, column=5, value=data['total_clicks'])
                ws.cell(row=row, column=6, value=data['ctr'])
                ws.cell(row=row, column=7, value=data['total_cost'])
        
        # Totals row
        total_row = header_row + (len(report_data) if report_data else 5) + 1
        ws.cell(row=total_row, column=1, value='Tổng:')
        ws.merge_cells(f'A{total_row}:B{total_row}')
        
        if report_data:
            total_records = len(report_data)
            total_views = sum(data['total_views'] for data in report_data)
            total_clicks = sum(data['total_clicks'] for data in report_data)
            total_cost = sum(data['total_cost'] for data in report_data)
        else:
            total_records = 0
            total_views = 0
            total_clicks = 0
            total_cost = 0

        ws.cell(row=total_row, column=3, value=total_records)
        ws.cell(row=total_row, column=4, value=total_views)
        ws.cell(row=total_row, column=5, value=total_clicks)
        ws.cell(row=total_row, column=7, value=total_cost)
        
        # Style totals row
        for col in range(1, 8):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        return wb
