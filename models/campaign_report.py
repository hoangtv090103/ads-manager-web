from configs.db import get_db_connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
from collections import defaultdict
import calendar
import unicodedata

class CampaignReport:
    @staticmethod
    def get_report_data(campaign_ids, start_date=None, end_date=None):
        with get_db_connection() as conn:
            cursor = conn.cursor()
            query = '''
                WITH website_metrics AS (
                    SELECT 
                        w.website_id,
                        w.domain_website,
                        c.camp_id,
                        c.ten_chien_dich,
                        c.ngay_bat_dau,
                        c.ngay_ket_thuc,
                        SUM(ag.tong_chi_phi) as tong_chi_phi,
                        SUM(ag.luot_xem) as luot_xem,
                        SUM(ag.luot_nhan) as luot_nhan,
                        CASE 
                            WHEN SUM(ag.luot_xem) > 0 
                            THEN CAST(((SUM(ag.luot_nhan)::float / SUM(ag.luot_xem)) * 100) AS DECIMAL(10,2))
                            ELSE 0 
                        END as ctr,
                        CASE 
                            WHEN SUM(ag.luot_nhan) > 0 
                            THEN CAST((SUM(ag.tong_chi_phi)::float / SUM(ag.luot_nhan)) AS DECIMAL(10,2))
                            ELSE 0 
                        END as cpc,
                        CASE 
                            WHEN SUM(ag.luot_xem) > 0 
                            THEN CAST((SUM(ag.tong_chi_phi)::float / SUM(ag.luot_xem)) * 1000 AS DECIMAL(10,2))
                            ELSE 0 
                        END as cpm,
                        SUM(ag.so_luong_mua_hang) as so_luong_mua_hang,
                        CASE 
                            WHEN SUM(ag.luot_nhan) > 0 
                            THEN CAST((SUM(ag.so_luong_mua_hang)::float / SUM(ag.luot_nhan)) * 100 AS DECIMAL(10,2))
                            ELSE 0 
                        END as conversion_rate,
                        CASE 
                            WHEN SUM(ag.so_luong_mua_hang) > 0 
                            THEN CAST((SUM(ag.tong_chi_phi)::float / SUM(ag.so_luong_mua_hang)) AS DECIMAL(10,2))
                            ELSE 0 
                        END as cps,
                        SUM(ag.videoview3s) as videoview3s,
                        SUM(ag.videowatchesat25) as videowatchesat25,
                        SUM(ag.videowatchesat50) as videowatchesat50,
                        SUM(ag.videowatchesat75) as videowatchesat75,
                        SUM(ag.videowatchesat100) as videowatchesat100,
                        EXTRACT(MONTH FROM c.ngay_bat_dau) as month,
                        EXTRACT(YEAR FROM c.ngay_bat_dau) as year,
                        EXTRACT(WEEK FROM c.ngay_bat_dau) as week,
                        DATE(c.ngay_bat_dau) as date
                    FROM campaign c
                    LEFT JOIN ads_group ag ON ag.camp_id = c.camp_id
                    LEFT JOIN ads_group_website agw ON agw.ads_group_id = ag.ads_group_id
                    LEFT JOIN website w ON w.website_id = agw.website_id
                    WHERE c.camp_id = ANY(%s::integer[])
                    AND c.active = true
                    AND ag.active = true
                    AND agw.active = true
                    GROUP BY w.website_id, w.domain_website, c.camp_id, c.ten_chien_dich, 
                             c.ngay_bat_dau, c.ngay_ket_thuc
                )
                SELECT * FROM website_metrics
            '''
            params = [campaign_ids]

            if start_date and end_date:
                query += ' WHERE ngay_bat_dau >= %s AND ngay_ket_thuc <= %s'
                params.extend([start_date, end_date])

            query += ' ORDER BY domain_website, ngay_bat_dau'
            
            cursor.execute(query, params)
            return [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]

    @staticmethod
    def get_week_range(date):
        """Get start and end date of a week for a given date."""
        # Get the day of week (0 = Monday, 6 = Sunday)
        day_of_week = date.weekday()
        
        # Calculate start of week (Monday)
        start_of_week = date - timedelta(days=day_of_week)
        
        # Calculate end of week (Sunday)
        end_of_week = start_of_week + timedelta(days=6)
        
        return start_of_week, end_of_week

    @staticmethod
    def generate_excel_report(report_data, start_date, end_date):
        wb = Workbook()
        
        # Tạo các sheets
        overview_sheet = wb.active
        overview_sheet.title = "Bao cao tong quan chien dich"
        monthly_sheet = wb.create_sheet("Bao cao chi tiet chien dich theo thang")
        weekly_sheet = wb.create_sheet("Bao cao chien dich theo tuan")
        daily_sheet = wb.create_sheet("Bao cao chien dich theo ngay")
        website_sheet = wb.create_sheet("Bao cao chien dich theo website")

        # Định nghĩa styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def add_common_header(sheet):
            sheet['A1'] = 'Cong ty co phan Novaon Digital'
            sheet['A2'] = 'Dia chi: 94 Nguyen Chinh, Phuong Liet, Hoang Mai, Ha Noi'
            sheet['A3'] = 'Hotline: 024 011 5632'
            
            sheet['A6'] = f'Tu ngay {start_date} den ngay {end_date}'
            sheet.merge_cells('A6:G6')
            sheet['A6'].alignment = center_align

            for row in range(1, 4):
                cell = sheet[f'A{row}']
                cell.font = Font(bold=True if row == 1 else False)

        def setup_monthly_sheet():
            add_common_header(monthly_sheet)
            monthly_sheet['A5'] = 'Bao cao chi tiet chien dich theo thang'
            monthly_sheet.merge_cells('A5:G5')
            monthly_sheet['A5'].alignment = center_align
            monthly_sheet['A5'].font = header_font

            # Thêm thông tin về thời gian cập nhật
            monthly_sheet['A8'] = 'So lieu duoc cap nhat luc:'
            monthly_sheet['B8'] = datetime.now().strftime('hh:mm:ss dd/mm/yyyy')
            monthly_sheet['A8'].font = Font(italic=True)

            # Headers cho bảng
            headers = ['Thang', 'Ten chien dich', 'Chi phi (VND)', 'Luot xem', 'Luot nhan', 
                      'CTR(%)', 'CPC (VND)', 'SL mua hang', '% chuyen doi mua hang', 'CPS (VND)',
                      'Video view 3s', 'Video watches at 25%', 'Video watches at 50%', 
                      'Video watches at 75%', 'Video watches at 100%']

            # Organize data by month
            monthly_data = defaultdict(list)
            for item in report_data:
                month_key = f"{int(item['month'])}/{int(item['year'])}"
                monthly_data[month_key].append(item)

            current_row = 10
            for month_key in sorted(monthly_data.keys(), key=lambda x: datetime.strptime(x, "%m/%Y")):
                # Add month header
                for col, header in enumerate(headers, 1):
                    cell = monthly_sheet.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border
                
                current_row += 1
                month_data = monthly_data[month_key]
                
                # Add data for each campaign in this month
                for data in month_data:
                    monthly_sheet.cell(row=current_row, column=1, value=month_key)
                    monthly_sheet.cell(row=current_row, column=2, value=data['ten_chien_dich'])
                    monthly_sheet.cell(row=current_row, column=3, value=data['tong_chi_phi'])
                    monthly_sheet.cell(row=current_row, column=4, value=data['luot_xem'])
                    monthly_sheet.cell(row=current_row, column=5, value=data['luot_nhan'])
                    monthly_sheet.cell(row=current_row, column=6, value=data['ctr'])
                    monthly_sheet.cell(row=current_row, column=7, value=data['cpc'])
                    monthly_sheet.cell(row=current_row, column=8, value=data['so_luong_mua_hang'])
                    monthly_sheet.cell(row=current_row, column=9, value=data['conversion_rate'])
                    monthly_sheet.cell(row=current_row, column=10, value=data['cps'])
                    monthly_sheet.cell(row=current_row, column=11, value=data['videoview3s'])
                    monthly_sheet.cell(row=current_row, column=12, value=data['videowatchesat25'])
                    monthly_sheet.cell(row=current_row, column=13, value=data['videowatchesat50'])
                    monthly_sheet.cell(row=current_row, column=14, value=data['videowatchesat75'])
                    monthly_sheet.cell(row=current_row, column=15, value=data['videowatchesat100'])
                    
                    # Apply borders and alignment
                    for col in range(1, 16):
                        cell = monthly_sheet.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    
                    current_row += 1
                
                # Add monthly totals
                monthly_sheet.cell(row=current_row, column=1, value=f"Tong {month_key}")
                monthly_sheet.merge_cells(f'A{current_row}:B{current_row}')
                
                # Calculate and add totals
                totals = {
                    'tong_chi_phi': sum(d['tong_chi_phi'] for d in month_data),
                    'luot_xem': sum(d['luot_xem'] for d in month_data),
                    'luot_nhan': sum(d['luot_nhan'] for d in month_data),
                    'ctr': sum(d['ctr'] for d in month_data) / len(month_data),
                    'cpc': sum(d['cpc'] for d in month_data) / len(month_data),
                    'so_luong_mua_hang': sum(d['so_luong_mua_hang'] for d in month_data),
                    'conversion_rate': sum(d['conversion_rate'] for d in month_data) / len(month_data),
                    'cps': sum(d['cps'] for d in month_data) / len(month_data),
                    'videoview3s': sum(d['videoview3s'] for d in month_data),
                    'videowatchesat25': sum(d['videowatchesat25'] for d in month_data),
                    'videowatchesat50': sum(d['videowatchesat50'] for d in month_data),
                    'videowatchesat75': sum(d['videowatchesat75'] for d in month_data),
                    'videowatchesat100': sum(d['videowatchesat100'] for d in month_data)
                }
                
                col = 3
                for key in ['tong_chi_phi', 'luot_xem', 'luot_nhan', 'ctr', 'cpc', 
                           'so_luong_mua_hang', 'conversion_rate', 'cps', 'videoview3s',
                           'videowatchesat25', 'videowatchesat50', 'videowatchesat75', 
                           'videowatchesat100']:
                    cell = monthly_sheet.cell(row=current_row, column=col)
                    cell.value = totals[key]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align
                    col += 1
                
                current_row += 2  # Add space between months

        def setup_weekly_sheet():
            add_common_header(weekly_sheet)
            weekly_sheet['A5'] = 'Bao cao chien dich theo tuan'
            weekly_sheet.merge_cells('A5:G5')
            weekly_sheet['A5'].alignment = center_align
            weekly_sheet['A5'].font = header_font

            # Thêm thông tin về thời gian cập nhật
            weekly_sheet['A8'] = 'So lieu duoc cap nhat luc:'
            weekly_sheet['B8'] = datetime.now().strftime('hh:mm:ss dd/mm/yyyy')
            weekly_sheet['A8'].font = Font(italic=True)

            # Headers cho bảng
            headers = ['Tuan', 'Ten chien dich', 'Chi phi (VND)', 'Luot xem', 'Luot nhan', 
                      'CTR(%)', 'CPC (VND)', 'SL mua hang', '% chuyen doi mua hang', 'CPS (VND)',
                      'Video view 3s', 'Video watches at 25%', 'Video watches at 50%', 
                      'Video watches at 75%', 'Video watches at 100%']

            # Organize data by week
            weekly_data = defaultdict(list)
            for item in report_data:
                start_date = datetime.strptime(str(item['ngay_bat_dau']), '%Y-%m-%d %H:%M:%S%z')
                week_start, week_end = CampaignReport.get_week_range(start_date)
                week_key = f"Tuan {int(item['week'])} ({week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')})"
                weekly_data[week_key].append(item)

            current_row = 10
            for week_key in sorted(weekly_data.keys()):
                # Add week header
                for col, header in enumerate(headers, 1):
                    cell = weekly_sheet.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border
                
                current_row += 1
                week_data = weekly_data[week_key]
                
                # Add data for each campaign in this week
                for data in week_data:
                    weekly_sheet.cell(row=current_row, column=1, value=week_key)
                    weekly_sheet.cell(row=current_row, column=2, value=data['ten_chien_dich'])
                    weekly_sheet.cell(row=current_row, column=3, value=data['tong_chi_phi'])
                    weekly_sheet.cell(row=current_row, column=4, value=data['luot_xem'])
                    weekly_sheet.cell(row=current_row, column=5, value=data['luot_nhan'])
                    weekly_sheet.cell(row=current_row, column=6, value=data['ctr'])
                    weekly_sheet.cell(row=current_row, column=7, value=data['cpc'])
                    weekly_sheet.cell(row=current_row, column=8, value=data['so_luong_mua_hang'])
                    weekly_sheet.cell(row=current_row, column=9, value=data['conversion_rate'])
                    weekly_sheet.cell(row=current_row, column=10, value=data['cps'])
                    weekly_sheet.cell(row=current_row, column=11, value=data['videoview3s'])
                    weekly_sheet.cell(row=current_row, column=12, value=data['videowatchesat25'])
                    weekly_sheet.cell(row=current_row, column=13, value=data['videowatchesat50'])
                    weekly_sheet.cell(row=current_row, column=14, value=data['videowatchesat75'])
                    weekly_sheet.cell(row=current_row, column=15, value=data['videowatchesat100'])
                    
                    # Apply borders and alignment
                    for col in range(1, 16):
                        cell = weekly_sheet.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    
                    current_row += 1
                
                # Add weekly totals
                weekly_sheet.cell(row=current_row, column=1, value=f"Tong {week_key}")
                weekly_sheet.merge_cells(f'A{current_row}:B{current_row}')
                
                # Calculate and add totals
                totals = {
                    'tong_chi_phi': sum(d['tong_chi_phi'] for d in week_data),
                    'luot_xem': sum(d['luot_xem'] for d in week_data),
                    'luot_nhan': sum(d['luot_nhan'] for d in week_data),
                    'ctr': sum(d['ctr'] for d in week_data) / len(week_data),
                    'cpc': sum(d['cpc'] for d in week_data) / len(week_data),
                    'so_luong_mua_hang': sum(d['so_luong_mua_hang'] for d in week_data),
                    'conversion_rate': sum(d['conversion_rate'] for d in week_data) / len(week_data),
                    'cps': sum(d['cps'] for d in week_data) / len(week_data),
                    'videoview3s': sum(d['videoview3s'] for d in week_data),
                    'videowatchesat25': sum(d['videowatchesat25'] for d in week_data),
                    'videowatchesat50': sum(d['videowatchesat50'] for d in week_data),
                    'videowatchesat75': sum(d['videowatchesat75'] for d in week_data),
                    'videowatchesat100': sum(d['videowatchesat100'] for d in week_data)
                }
                
                col = 3
                for key in ['tong_chi_phi', 'luot_xem', 'luot_nhan', 'ctr', 'cpc', 
                           'so_luong_mua_hang', 'conversion_rate', 'cps', 'videoview3s',
                           'videowatchesat25', 'videowatchesat50', 'videowatchesat75', 
                           'videowatchesat100']:
                    cell = weekly_sheet.cell(row=current_row, column=col)
                    cell.value = totals[key]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align
                    col += 1
                
                current_row += 2  # Add space between weeks

        def setup_daily_sheet():
            add_common_header(daily_sheet)
            daily_sheet['A5'] = 'Bao cao chien dich theo ngay'
            daily_sheet.merge_cells('A5:G5')
            daily_sheet['A5'].alignment = center_align
            daily_sheet['A5'].font = header_font

            # Thêm thông tin về thời gian cập nhật
            daily_sheet['A8'] = 'So lieu duoc cap nhat luc:'
            daily_sheet['B8'] = datetime.now().strftime('hh:mm:ss dd/mm/yyyy')
            daily_sheet['A8'].font = Font(italic=True)

            # Headers cho bảng
            headers = ['Ngay hien thi lieu', 'Ten chien dich', 'Chi phi (VND)', 'Luot xem', 'Luot nhan', 
                      'CTR(%)', 'CPC (VND)', 'SL mua hang', '% chuyen doi mua hang', 'CPS (VND)',
                      'Video view 3s', 'Video watches at 25%', 'Video watches at 50%', 
                      'Video watches at 75%', 'Video watches at 100%']

            # Organize data by date
            daily_data = defaultdict(list)
            for item in report_data:
                date_key = item['date'].strftime('%d/%m/%Y') if item['date'] else 'N/A'
                daily_data[date_key].append(item)

            current_row = 10
            for date_key in sorted(daily_data.keys(), key=lambda x: datetime.strptime(x, '%d/%m/%Y') if x != 'N/A' else datetime.max):
                # Add date header
                for col, header in enumerate(headers, 1):
                    cell = daily_sheet.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border
                
                current_row += 1
                date_data = daily_data[date_key]
                
                # Add data for each campaign in this date
                for data in date_data:
                    daily_sheet.cell(row=current_row, column=1, value=date_key)
                    daily_sheet.cell(row=current_row, column=2, value=data['ten_chien_dich'])
                    daily_sheet.cell(row=current_row, column=3, value=data['tong_chi_phi'])
                    daily_sheet.cell(row=current_row, column=4, value=data['luot_xem'])
                    daily_sheet.cell(row=current_row, column=5, value=data['luot_nhan'])
                    daily_sheet.cell(row=current_row, column=6, value=data['ctr'])
                    daily_sheet.cell(row=current_row, column=7, value=data['cpc'])
                    daily_sheet.cell(row=current_row, column=8, value=data['so_luong_mua_hang'])
                    daily_sheet.cell(row=current_row, column=9, value=data['conversion_rate'])
                    daily_sheet.cell(row=current_row, column=10, value=data['cps'])
                    daily_sheet.cell(row=current_row, column=11, value=data['videoview3s'])
                    daily_sheet.cell(row=current_row, column=12, value=data['videowatchesat25'])
                    daily_sheet.cell(row=current_row, column=13, value=data['videowatchesat50'])
                    daily_sheet.cell(row=current_row, column=14, value=data['videowatchesat75'])
                    daily_sheet.cell(row=current_row, column=15, value=data['videowatchesat100'])
                    
                    # Apply borders and alignment
                    for col in range(1, 16):
                        cell = daily_sheet.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    
                    current_row += 1
                
                # Add daily totals
                daily_sheet.cell(row=current_row, column=1, value=f"Tong {date_key}")
                daily_sheet.merge_cells(f'A{current_row}:B{current_row}')
                
                # Calculate and add totals
                totals = {
                    'tong_chi_phi': sum(d['tong_chi_phi'] for d in date_data),
                    'luot_xem': sum(d['luot_xem'] for d in date_data),
                    'luot_nhan': sum(d['luot_nhan'] for d in date_data),
                    'ctr': sum(d['ctr'] for d in date_data) / len(date_data),
                    'cpc': sum(d['cpc'] for d in date_data) / len(date_data),
                    'so_luong_mua_hang': sum(d['so_luong_mua_hang'] for d in date_data),
                    'conversion_rate': sum(d['conversion_rate'] for d in date_data) / len(date_data),
                    'cps': sum(d['cps'] for d in date_data) / len(date_data),
                    'videoview3s': sum(d['videoview3s'] for d in date_data),
                    'videowatchesat25': sum(d['videowatchesat25'] for d in date_data),
                    'videowatchesat50': sum(d['videowatchesat50'] for d in date_data),
                    'videowatchesat75': sum(d['videowatchesat75'] for d in date_data),
                    'videowatchesat100': sum(d['videowatchesat100'] for d in date_data)
                }
                
                col = 3
                for key in ['tong_chi_phi', 'luot_xem', 'luot_nhan', 'ctr', 'cpc', 
                           'so_luong_mua_hang', 'conversion_rate', 'cps', 'videoview3s',
                           'videowatchesat25', 'videowatchesat50', 'videowatchesat75', 
                           'videowatchesat100']:
                    cell = daily_sheet.cell(row=current_row, column=col)
                    cell.value = totals[key]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align
                    col += 1
                
                current_row += 2  # Add space between dates

        def setup_website_sheet():
            add_common_header(website_sheet)
            website_sheet['A5'] = 'Bao cao chien dich theo website'
            website_sheet.merge_cells('A5:G5')
            website_sheet['A5'].alignment = center_align
            website_sheet['A5'].font = header_font

            # Thêm thông tin về thời gian cập nhật
            website_sheet['A8'] = 'So lieu duoc cap nhat luc:'
            website_sheet['B8'] = datetime.now().strftime('hh:mm:ss dd/mm/yyyy')
            website_sheet['A8'].font = Font(italic=True)

            # Headers cho bảng
            headers = ['Website', 'Ten chien dich', 'Chi phi (VND)', 'Luot xem', 'Luot nhan', 
                      'CTR(%)', 'CPC (VND)', 'SL mua hang', '% chuyen doi mua hang', 'CPS (VND)',
                      'Video view 3s', 'Video watches at 25%', 'Video watches at 50%', 
                      'Video watches at 75%', 'Video watches at 100%']

            # Organize data by website
            website_data = defaultdict(list)
            for item in report_data:
                website_key = item['domain_website'] if item['domain_website'] else 'N/A'
                website_data[website_key].append(item)

            current_row = 10
            for website_key in sorted(website_data.keys()):
                # Add website header
                for col, header in enumerate(headers, 1):
                    cell = website_sheet.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border
                
                current_row += 1
                website_campaigns = website_data[website_key]
                
                # Add data for each campaign in this website
                for data in website_campaigns:
                    website_sheet.cell(row=current_row, column=1, value=website_key)
                    website_sheet.cell(row=current_row, column=2, value=data['ten_chien_dich'])
                    website_sheet.cell(row=current_row, column=3, value=data['tong_chi_phi'])
                    website_sheet.cell(row=current_row, column=4, value=data['luot_xem'])
                    website_sheet.cell(row=current_row, column=5, value=data['luot_nhan'])
                    website_sheet.cell(row=current_row, column=6, value=data['ctr'])
                    website_sheet.cell(row=current_row, column=7, value=data['cpc'])
                    website_sheet.cell(row=current_row, column=8, value=data['so_luong_mua_hang'])
                    website_sheet.cell(row=current_row, column=9, value=data['conversion_rate'])
                    website_sheet.cell(row=current_row, column=10, value=data['cps'])
                    website_sheet.cell(row=current_row, column=11, value=data['videoview3s'])
                    website_sheet.cell(row=current_row, column=12, value=data['videowatchesat25'])
                    website_sheet.cell(row=current_row, column=13, value=data['videowatchesat50'])
                    website_sheet.cell(row=current_row, column=14, value=data['videowatchesat75'])
                    website_sheet.cell(row=current_row, column=15, value=data['videowatchesat100'])
                    
                    # Apply borders and alignment
                    for col in range(1, 16):
                        cell = website_sheet.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    
                    current_row += 1
                
                # Add website totals
                website_sheet.cell(row=current_row, column=1, value=f"Tổng {website_key}")
                website_sheet.merge_cells(f'A{current_row}:B{current_row}')
                
                # Calculate and add totals
                totals = {
                    'tong_chi_phi': sum(d['tong_chi_phi'] for d in website_campaigns),
                    'luot_xem': sum(d['luot_xem'] for d in website_campaigns),
                    'luot_nhan': sum(d['luot_nhan'] for d in website_campaigns),
                    'ctr': sum(d['ctr'] for d in website_campaigns) / len(website_campaigns),
                    'cpc': sum(d['cpc'] for d in website_campaigns) / len(website_campaigns),
                    'so_luong_mua_hang': sum(d['so_luong_mua_hang'] for d in website_campaigns),
                    'conversion_rate': sum(d['conversion_rate'] for d in website_campaigns) / len(website_campaigns),
                    'cps': sum(d['cps'] for d in website_campaigns) / len(website_campaigns),
                    'videoview3s': sum(d['videoview3s'] for d in website_campaigns),
                    'videowatchesat25': sum(d['videowatchesat25'] for d in website_campaigns),
                    'videowatchesat50': sum(d['videowatchesat50'] for d in website_campaigns),
                    'videowatchesat75': sum(d['videowatchesat75'] for d in website_campaigns),
                    'videowatchesat100': sum(d['videowatchesat100'] for d in website_campaigns)
                }
                
                col = 3
                for key in ['tong_chi_phi', 'luot_xem', 'luot_nhan', 'ctr', 'cpc', 
                           'so_luong_mua_hang', 'conversion_rate', 'cps', 'videoview3s',
                           'videowatchesat25', 'videowatchesat50', 'videowatchesat75', 
                           'videowatchesat100']:
                    cell = website_sheet.cell(row=current_row, column=col)
                    cell.value = totals[key]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_align
                    col += 1
                
                current_row += 2  # Add space between websites

        # Thực hiện setup các sheets
        setup_monthly_sheet()
        setup_weekly_sheet()
        setup_daily_sheet()
        setup_website_sheet()

        # Điều chỉnh độ rộng cột cho tất cả các sheets
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

        return wb